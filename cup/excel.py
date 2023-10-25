from openpyxl import Workbook, load_workbook

class ExcelFile:
    def __init__(self, file_name):
        self.file_name = file_name
    
    def write_cells(self, page_name, matrix):
        excel_file = load_workbook(self.file_name)
        page = excel_file[page_name]
        ids = [cell.value for cell in page["A"]]
        for lst in matrix:
            if lst[0] not in ids:
                page.append(lst)
        excel_file.save(self.file_name)
    
    def get_messages(self):
        excel_file = load_workbook(self.file_name)
        page = excel_file["messages"]
        messages = {} # {"214235": {"text": "hello", "status": ""}}
        for row in page:
            if row[0].row == 1:
                continue
            messages[row[0].value] = {"text": row[1].value, "status": row[2].value}
        return messages
    
    def update_messages(self, messages):
        excel_file = load_workbook(self.file_name)
        page = excel_file["messages"]
        for row in page:
            if row[0].row == 1:
                continue
            row[2].value = messages[row[0].value]["status"]
        excel_file.save(self.file_name)
    
    def get_cup_names(self):
        excel_file = load_workbook(self.file_name)
        page = excel_file["cups"]
        names = ""
        for row in page:
            if row[0].row == 1:
                continue
            names += row[1].value
            names += "\n" 
        return names